from collections import defaultdict
import pandas as pd
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook
from PIL import Image, ExifTags
import streamlit as st
import os


# environment variables
os.environ["ENV"] = st.secrets["aws"]["ENV"]


# Image utilities
def handle_image_orientation(image: Image.Image) -> Image.Image:
    """
    Handle image orientation based on EXIF data.
    
    Args:
        image: PIL Image object
        
    Returns:
        PIL Image object with correct orientation
    """
    try:
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == 'Orientation':
                break
        exif = dict(image._getexif().items())

        if exif[orientation] == 3:
            image = image.rotate(180, expand=True)
        elif exif[orientation] == 6:
            image = image.rotate(270, expand=True)
        elif exif[orientation] == 8:
            image = image.rotate(90, expand=True)
    except (AttributeError, KeyError, IndexError):
        print("Image does not have exif data")
        # cases: image don't have exif data
        pass
    
    return image

# substrate analysis
def generate_keys(key_list, multiplier = 3):
    new_list = []

    for label in key_list:
        new_list.extend([label]*multiplier)

    return new_list


def create_substrate_dataframe(response_data: dict, csv_name: str) -> pd.DataFrame:
    segment_distances = ["0 - 19.5m", "25 - 44.5m", "50 - 65.5m", "75 - 94.5m"]

    # pop out the slate info 
    info_segment = response_data.pop('info_segment', None)
    # get unique keys
    response_keys = list(response_data.keys())
    # create dataframes
    df = pd.concat([pd.DataFrame.from_dict(response_data[key]) for key in list(response_data.keys())], axis = 1)

    # create unique column names
    column_names = []
    for num in range(len(list(response_data.keys()))):
        column_names.extend([f"distance_{num}", f"label_{num}", f"clear_{num}"])
    # set column names
    df.columns = column_names

    # generate multi-level index
    segment_list = generate_keys(list(response_data.keys()))
    merge_list = generate_keys(segment_distances)

    # create multi index
    arrays = [
        segment_list,
        merge_list,
        column_names
    ]
    columns = pd.MultiIndex.from_arrays(arrays)
    # create a dataframe
    df = pd.DataFrame(df.iloc[:,:].values, columns=columns)  

    # save the dataframe to a csv file
    df.to_csv(csv_name, index=False)

    return df, info_segment


def extract_details(info: dict) -> list:

    return [info["distance"], info["label"], info["label_status"]]


def extract_single_attributes(selected_set: list, index_val: int) -> list:
    sub_segments = []
    # first segment
    first_set = selected_set[index_val]
    second_set = selected_set[index_val + 20]
    sub_segments.extend(extract_details(first_set))
    sub_segments.extend(extract_details(second_set))

    return sub_segments


def substrate_excel_creation(response_data: dict, info_data: dict, excel_name: str):
    information_data = info_data
    print(f"========={info_data}===========")
    selected_set_one = response_data["segment_one"]
    selected_set_two = response_data["segment_two"]
    selected_set_three = response_data["segment_three"]
    selected_set_four = response_data["segment_four"]


    final_segments = []
    for index in range(20):
        sub_set_segments = []
        # first segment
        sub_set_segments.extend(extract_single_attributes(selected_set_one, index))
        # seconds segment
        sub_set_segments.extend(extract_single_attributes(selected_set_two, index))
        # seconds segment
        sub_set_segments.extend(extract_single_attributes(selected_set_three, index))
        # seconds segment
        sub_set_segments.extend(extract_single_attributes(selected_set_four, index))
        # append the segment
        final_segments.append(sub_set_segments)

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(excel_name)
    worksheet = workbook.add_worksheet()

    # Text Formats
    text_bold = workbook.add_format({'bold': True, 'font_name': 'Arial'})
    text_normal_bold = workbook.add_format({'bold': True, 'center_across': True, 'font_name': 'Arial'})
    text_bold_red = workbook.add_format({'bold': True, 'font_color': 'red', 'font_name': 'Arial'})

    # Add a bold format to use to highlight cells.
    main_bold = workbook.add_format({'bold': True, 'center_across': False, 'border': True, 'font_name': 'Arial'})
    bold_red = workbook.add_format({'bold': True, 'font_color': 'red', 'font_name': 'Arial'})

    # Table Cell Formats
    cell_bold = workbook.add_format({'bold': True, 'center_across': True, 'border': True, 'font_name': 'Arial'})
    cell_normal = workbook.add_format({'center_across': True, 'border': True, 'font_name': 'Arial'})
    cell_normal_red = workbook.add_format({'center_across': True, 'border': True, 'font_color': 'red', 'font_name': 'Arial'})

    # Add a border
    bold_border = workbook.add_format({'bold': True, 'border': True,'font_name': 'Arial', 'bg_color': '#29E37D'})
    border = workbook.add_format({'border': True, 'font_name': 'Arial'})

    # Add a number format for cells with highlight
    not_clear = workbook.add_format({'bold': True, 'bg_color': 'red', 'border': True, 'font_name': 'Arial'})

    # Instructions
    instructions =  workbook.add_format({'font_name': 'Arial', 'bold': True, 'font_color': 'red','align': 'left'})

    # Lower Table
    lower_table_cell_border = workbook.add_format({ 'border': 2, 'bold': True,'font_name': 'Arial'})
    lower_table_cell_normal = workbook.add_format({ 'font_name': 'Arial'})

    lower_table_left_border_normal = workbook.add_format({'right': 2, 'font_name': 'Arial', 'align': 'left'})

    no_left_border = workbook.add_format({'top': 2,'right': 2,'bottom': 2,'font_name': 'Arial', 'bold': True, 'font_color': 'red', 'align': 'left'})
    hash_tag = workbook.add_format({'top': 2,'bottom': 2,'font_name': 'Arial', 'bold': True, 'align': 'left'})


    #Write info headers.
    worksheet.merge_range("A1:C1", "Site Name", main_bold)
    worksheet.merge_range("A2:C2", "Transcript Depth", main_bold)
    worksheet.merge_range("A3:C3", "Team Leader", main_bold)
    worksheet.merge_range("A4:C4", "Start Time", main_bold)

    worksheet.merge_range("I1:K1", "Country/island", main_bold)
    worksheet.merge_range("I2:K2", "Date", main_bold)
    worksheet.merge_range("I3:K3", "Data recorded by", main_bold)
    worksheet.merge_range("I4:K4", "Full Name", main_bold)
    worksheet.merge_range("L4:P4", "", main_bold)  

    worksheet.merge_range("A6:P6", "Enter the Substrate Codes into the white boxes below - just like your field sheet.", bold_red)

    bold = workbook.add_format({'bold': True, 'font_name': 'Arial'})
    normal = workbook.add_format({'font_name': 'Arial'})

    # ---- Left block ----
    worksheet.write('A7',  'HC',  bold);   worksheet.write('B7',  'hard coral',             normal)
    worksheet.write('A8',  'NIA', bold);   worksheet.write('B8',  'nutrient indicator algae',normal)
    worksheet.write('A9',  'RB',  bold);   worksheet.write('B9',  'rubble',                  normal)
    worksheet.write('A10', 'OT',  bold);   worksheet.write('B10', 'other',                   normal)

    # ---- Middle block ----
    worksheet.write('E7',  'SC',  bold);   worksheet.write('F7',  'soft coral', normal)
    worksheet.write('E8',  'SP',  bold);   worksheet.write('F8',  'sponge',     normal)
    worksheet.write('E9',  'SD',  bold);   worksheet.write('F9',  'sand',       normal)

    # ---- Right block ----
    worksheet.write('I7',  'RKC', bold);   worksheet.write('J7',  'recently killed coral', normal)
    worksheet.write('I8',  'RC',  bold);   worksheet.write('J8',  'rock',                  normal)
    worksheet.write('I9',  'SI',  bold);   worksheet.write('J9',  'silt/clay',              normal)

    
    # Write some data headers.
    worksheet.merge_range("A12:J12", 
                          "(For first segment, if start point is 0 m, last point is 19.5 m)", 
                          workbook.add_format({'italic': True, 'font_name': 'Arial'}))
    

    # Table 
    worksheet.merge_range("A13:D13", "Segment One", cell_bold)
    worksheet.merge_range("E13:H13", "Segment Two", cell_bold)
    worksheet.merge_range("I13:L13", "Segment Three", cell_bold)
    worksheet.merge_range("M13:P13", "Segment Four", cell_bold)

    # distances
    worksheet.merge_range("A14:D14", "0 - 19.5m", cell_bold)
    worksheet.merge_range("E14:H14", "25 - 44.5m", cell_bold)
    worksheet.merge_range("I14:L14", "50 - 69.5m", cell_bold)
    worksheet.merge_range("M14:P14", "75 - 94.5", cell_bold)

    info_col = 3
    info_row = 0

    info_fields = [
        "site_name",
        "country_island",
        "depth",
        "date",
        "team_leader",
        "data_recorded_by",
        "time",
    ]

    for i in range(0, len(info_fields), 2):
        worksheet.merge_range(
            info_row,        # row index (0-based)
            info_col,
            info_row,
            info_col + 4,
            info_data.get(info_fields[i], ""),
            border
        )

        if i + 1 < len(info_fields):
            worksheet.merge_range(
                info_row,        # row index (0-based)
                info_col+8,
                info_row,
                info_col + 12,
                info_data.get(info_fields[i+1], ""),
                border
            )
        info_row += 1
            

    
    row = 14
    # adding records
    for diff_segments in final_segments:
        col = 0
        for range_index in range(0, 24, 3):
            worksheet.write(row, col, diff_segments[range_index], bold_border)
            col += 1
            worksheet.write(row, col, diff_segments[range_index +1], not_clear if not diff_segments[range_index +2] else border)
            col += 1
        row += 1


    worksheet.write(35, 0, "# of HC with disease:", text_bold)
    worksheet.write(36, 0, "# of HC with bleaching:", text_bold)
    worksheet.write(37, 0, "% of HC with disease:", text_bold_red)
    worksheet.write(38, 0, "% of HC with bleaching:", text_bold_red)
    worksheet.write(39, 0, "% of RKC", text_bold_red)

    worksheet.merge_range("F35:G35", "S1", text_normal_bold)
    worksheet.merge_range("H35:I35", "S2", text_normal_bold)
    worksheet.merge_range("J35:K35", "S3", text_normal_bold)
    worksheet.merge_range("L35:M35", "S4", text_normal_bold)
    worksheet.merge_range("N35:O35", "Segment Mean", text_normal_bold)

    worksheet.merge_range("F36:G36", "0", cell_normal)
    worksheet.merge_range("H36:I36", "0", cell_normal)
    worksheet.merge_range("J36:K36", "0", cell_normal)
    worksheet.merge_range("L36:M36", "0", cell_normal)
    worksheet.merge_range("N36:O36", "=(F36+H36+J36+L36)/4", cell_normal_red)

    worksheet.merge_range("F37:G37", "0", cell_normal)
    worksheet.merge_range("H37:I37", "0", cell_normal)
    worksheet.merge_range("J37:K37", "0", cell_normal)
    worksheet.merge_range("L37:M37", "0", cell_normal)
    worksheet.merge_range("N37:O37", '=(F37+H37+J37+L37)/4', cell_normal_red)

    worksheet.merge_range("F38:G38", '=IF(B52=0,0,F36/B52)', cell_normal_red)
    worksheet.merge_range("H38:I38", '=IF(D52=0,"0",H36/D52)', cell_normal_red)
    worksheet.merge_range("J38:K38", '=IF(F52=0,"0",J36/F52)', cell_normal_red)
    worksheet.merge_range("L38:M38", '=IF(H52=0,"0",L36/H52)', cell_normal_red)
    worksheet.merge_range("N38:O38", '=(F38+H38+J38+L38)/4', cell_normal_red)

    worksheet.merge_range("F39:G39", '=IF(B52=0,0,F37/B52)', cell_normal_red)
    worksheet.merge_range("H39:I39", '=IF(D52=0,"0",H37/D52)', cell_normal_red)
    worksheet.merge_range("J39:K39", '=IF(F52=0,"0",J37/F52)', cell_normal_red)
    worksheet.merge_range("L39:M39", '=IF(H52=0,"0",L37/H52)', cell_normal_red)
    worksheet.merge_range("N39:O39", '=(F39+H39+J39+L39)/4', cell_normal_red)

    worksheet.merge_range("F40:G40", '=B54/40', cell_normal_red)
    worksheet.merge_range("H40:I40", '=D54/40', cell_normal_red)
    worksheet.merge_range("J40:K40", '=F54/40', cell_normal_red)
    worksheet.merge_range("L40:M40", '=H54/40', cell_normal_red)
    worksheet.merge_range("N40:O40", '=(F40+H40+J40+L40)/4', cell_normal_red)

    worksheet.merge_range("A42:E42", "If mean RKC is > 10%, is the primary cause:", text_bold)
    worksheet.merge_range("G42:N42", "", workbook.add_format({'font_name': 'Arial', 'bg_color': '#969696'}))

    worksheet.merge_range("A45:B45", "Comments:", text_bold)
    worksheet.merge_range("C45:P45", "", workbook.add_format({'font_name': 'Arial', 'bg_color': '#969696'}))
    worksheet.merge_range("C46:P46", "", workbook.add_format({'font_name': 'Arial', 'bg_color': '#969696'}))


    worksheet.write(47, 0, "Reef Check Summary Data", workbook.add_format({'bold': True,'underline': True, 'font_name': 'Arial'}))
    worksheet.merge_range("A49:J50", "DO NOT TYPE DATA BELOW THIS LINE", text_bold_red)

    # Summary Data Table-1
    worksheet.merge_range("A51:B51", "Total S1", lower_table_cell_border)
    worksheet.merge_range("C51:D51", "Total S2", lower_table_cell_border)
    worksheet.merge_range("E51:F51", "Total S3", lower_table_cell_border)
    worksheet.merge_range("G51:H51", "Total S4", lower_table_cell_border)
    worksheet.merge_range("I51:J51", "Grand total", lower_table_cell_border)


    subs_list_ = ["HC", "SC", "RKC", "NIA", "SP", "RC", "RB", "SD", "SI", "OT"]

    rows_1 = 51
    for sub in subs_list_:
        for k in range(0,10,2):
            worksheet.write(rows_1, k, f"{sub}", lower_table_cell_normal)
        rows_1 += 1


    rows_2 = 51
    for sub in subs_list_:
        worksheet.write(rows_2, 1, f'=COUNTIF(B15:B34:D15:D34,"{sub}")', lower_table_left_border_normal)
        worksheet.write(rows_2, 3, f'=COUNTIF(F15:F34:H15:H34,"{sub}")', lower_table_left_border_normal)
        worksheet.write(rows_2, 5, f'=COUNTIF(J15:J34:L15:L34,"{sub}")', lower_table_left_border_normal)
        worksheet.write(rows_2, 7, f'=COUNTIF(N15:N34:P15:P34,"{sub}")', lower_table_left_border_normal)
        worksheet.write(rows_2, 9, f"=(B{rows_2+1}+D{rows_2+1}+F{rows_2+1}+H{rows_2+1})", lower_table_left_border_normal)
        rows_2 += 1

    
    column_letters = ['B', 'D', 'F', 'H', 'J']
    t = 0
    for m in range(0,10,2):
        worksheet.write(61, m, '#', hash_tag)
        worksheet.write(61, m+1, f'={column_letters[t]}52+{column_letters[t]}53+{column_letters[t]}54+{column_letters[t]}55+{column_letters[t]}56+{column_letters[t]}57+{column_letters[t]}58+{column_letters[t]}59+{column_letters[t]}60+{column_letters[t]}61', no_left_border)
        t += 1

    # Lower second table
    lower_table_cell_border_2 = workbook.add_format({ 'border': 2, 'bold': True,'font_name': 'Arial', 'font_size': 10, 'align': 'center'})
    

    rows_3 = 51
    for sub in subs_list_:
        worksheet.write(rows_3, 10, f"{sub}", workbook.add_format({ 'font_name': 'Arial', 'align': 'right'}))
        rows_3 += 1

    worksheet.merge_range("L50:L51", "Mean count", lower_table_cell_border_2)
    worksheet.merge_range("M50:N51", "Mean % per segment", lower_table_cell_border_2)
    worksheet.merge_range("O50:O51", "SD", lower_table_cell_border_2)
    worksheet.merge_range("P50:P51", "%     S1", lower_table_cell_border_2)
    worksheet.merge_range("Q50:Q51", "%     S2", lower_table_cell_border_2)
    worksheet.merge_range("R50:R51", "%     S3", lower_table_cell_border_2)
    worksheet.merge_range("S50:S51", "%     S4", lower_table_cell_border_2)


    LT_format_1 = workbook.add_format({ 'left': 2,'right': 2, 'font_name': 'Arial', 'align': 'center', 'font_color': 'red'})
    LT_format_2 = workbook.add_format({ 'left': 2,'right': 2, 'num_format': '0%', 'font_name': 'Arial', 'font_size': 10, 'align': 'center', 'font_color': 'red'})


    for len_ in range(51, 61):
        worksheet.write(len_, 11, f"=J{len_+1}/4", LT_format_1)
        worksheet.merge_range(f"M{len_+1}:N{len_+1}", f"=AVERAGE(P{len_+1}:S{len_+1})", LT_format_1)
        worksheet.write(len_, 14, f"=ROUND(STDEV(P{len_+1}:S{len_+1}), 4)", LT_format_1)  
        worksheet.write(len_, 15, f"=B{len_+1}/40", LT_format_2)
        worksheet.write(len_, 16, f"=D{len_+1}/40", LT_format_2)
        worksheet.write(len_, 17, f"=F{len_+1}/40", LT_format_2)
        worksheet.write(len_, 18, f"=H{len_+1}/40", LT_format_2)

    worksheet.merge_range("L62:S62", "", workbook.add_format({ 'top': 2}))

    # Final Instructions
    worksheet.write(62, 0, "TOTALS MUST = 40 FOR EACH SEGMENT",instructions)
    worksheet.write(64, 0, "FILL IN WHITE BOXES ONLY",instructions)
    worksheet.write(64, 11, "PLEASE TURN TO THE GRAPHS TAB",instructions)
    worksheet.write(66, 0, "This page is ready to print",instructions)


    workbook.close()


def load_and_prepare_excel_for_substrate(excel_name: str):
    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_name)
    with BytesIO() as buffer:
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def create_fish_slate_dataframe(response_data: dict, csv_name: str) -> pd.DataFrame:
    # distances are constants
    distances = ["0 - 20m", "25 - 45m", "50 - 75m", "75 - 95m"]

    main_keys = list(response_data.keys())

    info_df = pd.concat([pd.DataFrame.from_dict(response_data[key_]) for key_ in main_keys])
    
    new_columns = []
    new_columns.append("name")
    for distance_index in range(len(distances)):
        new_columns.extend([distances[distance_index], "set_{}_clear".format(distance_index)])

    info_df.columns = new_columns

    # save the dataframe to a csv file
    info_df.to_csv(csv_name, index=False)

    return info_df


def fish_slate_excel_creation(response_data: dict, info_data: dict, excel_name: str):

    data_list = []

    for key_ in list(response_data.keys()):
        data_list.extend(response_data[key_])


    distances = ["0 - 20m", "25 - 45m", "50 - 75m", "75 - 95m"]
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(excel_name)
    worksheet = workbook.add_worksheet()

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True, 'center_across': True, 'border': True})

    # sub category
    sub_format = workbook.add_format({'bold': True, 'center_across': True, 'border': True, 'bg_color': 'green'})

    # Add a border
    border = workbook.add_format({'border': True})

    # Add a number format for cells with money.
    not_clear = workbook.add_format({'bold': True, 'bg_color': 'red', 'border': True})

    #Write info headers.
    worksheet.merge_range("A1:B1", "Site Name", bold)
    worksheet.merge_range("C1:D1", "Country/ island", bold)
    worksheet.merge_range("E1:F1", "Team Leader", bold)
    worksheet.merge_range("G1:H1", "Data Recorded By", bold)
    worksheet.merge_range("I1:J1", "Depth", bold)
    worksheet.merge_range("K1:L1", "Date", bold)
    worksheet.merge_range("M1:N1", "Time", bold)
    
    
    
    worksheet.merge_range("A4:E4", "Fish Slate Analysis", bold)
    worksheet.write(4, 0, "Type", bold)
    worksheet.set_column('A:A', 30)
    
    # Write Fish Slate record data
    info_col = 0
    info_row = 2

    info_fields = [
        "site_name",
        "country_island",
        "team_leader",
        "data_recorded_by",
        "depth",
        "date",
        "time",
    ]

    for field in info_fields:
        worksheet.merge_range(
            info_row - 1,        # row index (0-based)
            info_col,
            info_row - 1,
            info_col + 1,
            info_data.get(field, ""),
            border
        )
        info_col += 2
    
    
    
    # distances
    for index in range(len(distances)):
        worksheet.write(4, index + 1, distances[index] , bold)

    row = 5
    col = 0
    for key_ in list(response_data.keys()):
        worksheet.write(row, col, key_, sub_format)
        row +=1 
        for data_dict in response_data[key_]:
            worksheet.write(row, col, data_dict["name"], border)
            worksheet.write(row, col+1, data_dict["distance_one"], not_clear if not data_dict["distance_one_clear"] else border)
            worksheet.write(row, col+2, data_dict["distance_two"], not_clear if not data_dict["distance_two_clear"] else border)
            worksheet.write(row, col+3, data_dict["distance_three"], not_clear if not data_dict["distance_three_clear"] else border)
            worksheet.write(row, col+4, data_dict["distance_four"], not_clear if not data_dict["distance_four_clear"] else border)
            row += 1


    workbook.close()


def load_and_prepare_excel_for_fish_slate(excel_name: str):
    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_name)
    with BytesIO() as buffer:
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def substrate_excel_data_extractor(data: pd.DataFrame) -> dict:
    suffixes = ["one", "two", "three", "four"]
    annots = defaultdict(list)
    column_list = list(data.columns)
    count = 0

    for index in range(0, 12, 3):
        distance = data[column_list[index]].to_list()
        label = data[column_list[index + 1]].to_list()
        status = data[column_list[index + 2]].to_list()

        for distance_, label_, status_ in zip(distance, label, status):
            annots["segment_{}".format(suffixes[count])].append({
                "distance": distance_,
                "label": label_,
                "label_status": status_
            })

        count +=1

    return dict(annots)


def extract_fish_details(fish_details: list) -> list:
    total_records = []

    for record_ in fish_details:
        sample_dict = {}
        sample_dict["name"] = record_["name"]
        sample_dict["distance_one"] = record_["0 - 20m"]
        sample_dict["distance_one_clear"] = record_["set_0_clear"]
        sample_dict["distance_two"] = record_["25 - 45m"]
        sample_dict["distance_two_clear"] = record_["set_1_clear"]
        sample_dict["distance_three"] = record_["50 - 75m"]
        sample_dict["distance_three_clear"] = record_["set_2_clear"]
        sample_dict["distance_four"] = record_["75 - 95m"]
        sample_dict["distance_four_clear"] = record_["set_3_clear"]
        total_records.append(sample_dict)

    return total_records


def fish_excel_data_extractor(data: pd.DataFrame) -> dict:
    # get data records
    records = data.to_dict(orient='records')

    annots = defaultdict(list)

    # original records
    fish_records = records[: 12]
    invertebrates_records = records[12: 26]
    impacts_records = records[26: 33]
    coral_disease_records = records[33: 35]
    rare_animals_records = records[35: 39]

    # processed records
    process_dict = {
        "fish": fish_records,
        "invertebrates": invertebrates_records,
        "impacts": impacts_records,
        "coral_disease": coral_disease_records,
        "rare_animals": rare_animals_records

    }
    
    for key_ in process_dict.keys():
        annots[key_].extend(extract_fish_details(process_dict[key_]))

    return dict(annots)